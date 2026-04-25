import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs" / "V0_progress.xlsx"
DATA_SOURCE = ROOT / "docs" / "V0_progress_data.json"


def style_sheet_header(ws, border, fill, font, align):
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def build_workbook():
    payload = json.loads(DATA_SOURCE.read_text(encoding="utf-8"))
    rows = payload["rows"]
    summary_data = payload["summary"]

    wb = Workbook()
    ws = wb.active
    ws.title = "V0_Status"

    headers = [
        "模块名称",
        "版本范围",
        "策划状态",
        "UI状态",
        "前端状态",
        "后端状态",
        "联调状态",
        "当前负责人",
        "风险/阻塞",
        "下一步动作",
        "最近更新时间",
    ]
    ws.append(headers)

    for row in rows:
        ws.append([row[header] for header in headers])

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    status_fills = {
        "未开始": "EDEDED",
        "进行中": "FFF2CC",
        "已完成": "C6E0B4",
        "待联调": "D9EAF7",
        "已联调": "9FD5B3",
        "阻塞": "F4CCCC",
        "暂不处理": "D9D2E9",
        "待确认": "FCE5CD",
        "不适用": "F3F3F3",
    }

    style_sheet_header(ws, border, header_fill, header_font, center)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = center if idx in {3, 4, 5, 6, 7, 11} else left
            if idx in {3, 4, 5, 6, 7}:
                fill = status_fills.get(str(cell.value))
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)

    widths = {1: 18, 2: 10, 3: 11, 4: 11, 5: 11, 6: 11, 7: 11, 8: 18, 9: 34, 10: 38, 11: 14}
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    legend = wb.create_sheet("Legend")
    legend.append(["状态值", "含义"])
    legend_rows = [
        ["未开始", "尚未进入实际制作或实现。"],
        ["进行中", "已经开始制作或实现，但尚未形成稳定结果。"],
        ["已完成", "该角色职责范围内的当前任务已经完成。"],
        ["待联调", "单侧已基本完成，但还需要和其他角色产物打通。"],
        ["已联调", "相关角色之间已经打通并完成验证。"],
        ["阻塞", "被明确问题卡住，不能继续推进。"],
        ["待确认", "存在需要拍板的问题，暂不能完全定案。"],
        ["暂不处理", "当前版本阶段内不作为优先实现项。"],
        ["不适用", "该角色当前不承担此项直接产出。"],
    ]
    for row in legend_rows:
        legend.append(row)
    style_sheet_header(legend, border, header_fill, header_font, center)
    for row in legend.iter_rows(min_row=2, max_row=legend.max_row):
        for idx, cell in enumerate(row, start=1):
            cell.border = border
            cell.alignment = left if idx == 2 else center
            if idx == 1:
                fill = status_fills.get(str(cell.value))
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
    legend.column_dimensions["A"].width = 12
    legend.column_dimensions["B"].width = 42
    legend.freeze_panes = "A2"

    summary = wb.create_sheet("Summary")
    summary_rows = [
        ["项目", "内容"],
        ["更新时间", summary_data["updated_at"]],
        ["当前说明", summary_data["description"]],
        ["当前重点", summary_data["current_focus"]],
        ["后续建议", summary_data["next_advice"]],
    ]
    for row in summary_rows:
        summary.append(row)
    style_sheet_header(summary, border, header_fill, header_font, center)
    for row in summary.iter_rows(min_row=2, max_row=summary.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = left
    summary.column_dimensions["A"].width = 14
    summary.column_dimensions["B"].width = 90

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT)
